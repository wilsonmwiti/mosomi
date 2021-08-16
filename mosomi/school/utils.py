import csv

from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook, load_workbook


def get_excel_content(file):
    """
    Save, Read an xlsx file and return its contents row by row
    """
    fs = FileSystemStorage()
    filename = fs.save(file.name, file)
    uploaded_file_url = fs.url(filename)

    extension = file.name.rsplit('.', 1)[1]
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        return convert_csv_to_xlsx(file_path)

    file_path = uploaded_file_url.split('/', 1)[1]
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    sheet = sheet_names[0]
    worksheet = workbook.get_sheet_by_name(sheet)

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    contacts = []

    for i in range(1, max_row + 1):
        person_details = []
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            person_details.append(cell_obj.value)
        contacts.append(person_details)

    headers = []
    for i in range(1, 2):
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            headers.append(cell_obj.value)
    context = {
        'contacts': contacts,
        'fields': headers,
        'file': file_path
    }

    return context


# convert csv to xlsx
def convert_csv_to_xlsx(file):
    """
    Method to convert uploaded .csv file to .xlsx before reading with openpyxl
     sample usage:
        file = convert_csv_to_xlsx(file_path)
    """
    f = file.rsplit('/', 1)[1]
    file_name = f.rsplit('.', 1)[0]
    wb = Workbook()
    ws = wb.active
    with open(file, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save('media/%s.xlsx' % file_name)

    uploaded_file_url = 'media/%s.xlsx' % file_name
    return get_excel_content_after_conversion(uploaded_file_url)



def get_excel_content_after_conversion(file_path):
    '''
    Save, Read an xlsx file and return its contents row by row
    '''
    workbook = load_workbook(file_path)

    sheet_names = workbook.sheetnames
    sheet = sheet_names[0]
    worksheet = workbook.get_sheet_by_name(sheet)

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    contacts = []

    for i in range(1, max_row + 1):
        person_details = []
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            person_details.append(cell_obj.value)
        contacts.append(person_details)

    headers = []
    for i in range(1, 2):
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            headers.append(cell_obj.value)
    context = {
        'contacts': contacts,
        'fields': headers,
        'file': file_path
    }
    return context


def get_message_parameters(message):
    '''
    This method returns all parameters passed in the message specified by a user
    sample usage
        parameters = get_message_parameters(message)
        eg:
        message = Hello [name], check your email [email] for more information
        print(get_message_parameters(message))

        output:
        ['name', 'email']
    '''
    string_list = message.split()
    parameters = []
    for string in string_list:
        correct_parameter = ''
        if string.endswith('.'):
            without_full_stop = string.strip().replace('.', '')
            if without_full_stop.endswith(']') and without_full_stop.startswith('['):
                l = string.replace('[', '')
                k = l.replace(']', '')
                correct_parameter = k

        elif string.endswith(','):
            without_comma = string.replace(',', '')
            if without_comma.endswith(']') and without_comma.startswith('['):
                l = string.replace('[', '')
                k = l.replace(']', '')
                correct_parameter = k
            # print(parameters)
        else:
            if string.endswith(']') and string.startswith('['):
                l = string.replace('[', '')
                k = l.replace(']', '')
                correct_parameter = k
            # print(parameters)
        if correct_parameter != '':
            if correct_parameter.endswith('.'):
                my_parameter = correct_parameter.replace('.', '')
                parameters.append(my_parameter)
            elif correct_parameter.endswith(','):
                my_parameter = correct_parameter.replace(',', '')
                parameters.append(my_parameter)
            else:
                my_parameter = correct_parameter
                parameters.append(my_parameter)
    return parameters


def get_parameter_column(parameters, worksheet):
    '''
    For each parameter returned on the get_message_parameters() method
    return the corresponding column in the uploaded excel file
    '''
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    parameter_cells = {}

    for parameter in parameters:
        for i in range(1, 2):
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)
                if cell_obj.value is not None:
                    if cell_obj.value.strip() == parameter.strip():
                        parameter_cells[parameter] = j
    return parameter_cells


def get_phone_number_column(phone_number_field, worksheet):
    '''
    This method returns the column corresponding to the phone number field specified by the user(client)
    '''
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    phone_number_column = ''
    for i in range(1, 2):
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            if cell_obj.value == phone_number_field:
                phone_number_column = j
    if phone_number_column != '':
        return phone_number_column