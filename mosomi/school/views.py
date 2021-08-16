import csv
import datetime
import os
from pathlib import Path
from pprint import pprint

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect

# Create your views here.
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook

from school.forms import *
from school.utils import get_excel_content, get_message_parameters, get_parameter_column, get_phone_number_column


class Merged:
    def __init__(self, phone_number, message):
        self.phone_number = phone_number
        self.message = message


@login_required()
def home(request):
    return render(request, 'school/school/dashboard.html')


def school_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        print(username, password)
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            messages.success(request, 'Login Successful')
            return redirect('School:home')
        else:
            messages.error(request, 'Invalid Email Or Password')
            return redirect('School:login')
    return render(request, 'school/registration/login.html')


def school_register(request):
    if request.method == 'POST':
        admin_form = SchoolAdminForm(request.POST)
        school_form = SchoolForm(request.POST)
        print(admin_form)
        print(school_form)
        if school_form.is_valid():
            new_school = school_form.save()
            if admin_form.is_valid():
                new_admin = admin_form.save(commit=False)
                new_admin.username = request.POST['email']
                new_admin.school = new_school
                new_admin.save()
                user = authenticate(username=request.POST['email'], password=request.POST['password2'])
                if user:
                    login(request, user)
                    messages.success(request, 'School Registered Successfully')
                    return redirect('School:home')
            else:
                form = SchoolAdminForm()
                context = {
                    'form': form
                }
                return render(request, 'school/registration/register.html', context)
        else:
            form = SchoolForm()
            context = {
                'form': form
            }
            return render(request, 'school/registration/register.html', context)
    return render(request, 'school/registration/register.html')


def school_logout(request):
    logout(request)
    messages.success(request, 'You have logged out successfully')
    return redirect('School:login')


@login_required()
def school_classes(request):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    classes = Form.objects.filter(school=school_admin.school)

    context = {
        'classes': classes
    }
    return render(request, 'school/school/school_classes.html', context)


@login_required()
def add_class(request):
    admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = admin.school
    # school =School.objects.get(id=school_id)
    if request.method == 'POST':
        Form.objects.create(
           name=request.POST['name'],
           school=school
        )
        return redirect("School:school_classes")
    context = {
        'school':school
    }
    return render(request, 'school/school/add_class.html', context)


@login_required()
def edit_class(request, class_id):
    form = Form.objects.get(id=class_id)
    if request.method == 'POST':
        form.name = request.POST['name']
        form.save()
        return redirect("School:school_classes")
    context = {
        'class': form
    }
    return render(request, 'school/school/edit_class.html', context)


@login_required()
def school_streams(request, form_id):
    streams = Stream.objects.filter(form_id=form_id)
    form = Form.objects.get(id=form_id)
    context = {
        'streams': streams,
        'form':form
    }
    return render(request, 'school/school/class_streams.html', context)


@login_required()
def add_stream(request, class_id):
    form = Form.objects.get(id=class_id)
    if request.method == 'POST':
        Stream.objects.create(
            name=request.POST['name'],
            form=form
        )
        messages.success(request, 'Stream Added Successfully')
        return redirect("School:school_streams", form.id)
    context = {
        'class': form
    }
    return render(request, 'school/school/add_stream.html', context)


@login_required()
def edit_stream(request, stream_id):
    stream = Stream.objects.get(id=stream_id)
    if request.method == 'POST':
        stream.name = request.POST['name']
        stream.save()
        messages.success(request, 'Stream Updated Successfully')
        return redirect("School:school_streams", stream.form.id)
    context = {
        'stream': stream
    }
    return render(request, 'school/school/edit_stream.html', context)


@login_required()
def school_students(request, stream_id):
    stream_students = StudentStream.objects.filter(stream_id=stream_id)
    ids = []
    for stream_student in stream_students:
        ids.append(stream_student.student.id)
    students = Student.objects.filter(id__in=list(set(ids)))
    print(students)
    context = {
        'students': students,
        'stream': Stream.objects.get(id=stream_id)
    }
    return render(request, 'school/school/stream_students.html', context)


@login_required()
def add_student(request, stream_id):
    stream = Stream.objects.get(id=stream_id)
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    forms = Form.objects.filter(school=school_admin.school)
    if request.method == 'POST':
        student = Student.objects.create(
            first_name=request.POST['first_name'],
            last_name=request.POST['last_name'],
            gender=request.POST['gender'],
            date_of_birth=request.POST['date_of_birth'],
            admission_number=request.POST['admission_number'],
            school=stream.form.school,
            guardian_id=request.POST['guardian'],
        )
        StudentStream.objects.create(
            stream=stream,
            student=student
        )
        StudentForm.objects.create(
            student=student,
            is_current=True,
            form_id=request.POST['school_class'],
            year=request.POST['year']
        )
        for subject in request.POST.getlist('subjects'):
            StudentFormSubject.objects.create(
                form_id=request.POST['school_class'],
                student=student,
                subject_id=subject
            )
        messages.success(request, 'Student Added Successfully')
        return redirect('School:school_students', stream.id)
    context = {
        'stream': stream,
        'fs': forms
    }
    pprint(context)
    return render(request, 'school/school/add_student.html', context)


@login_required()
def add_guardian(request, stream_id):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    forms = Form.objects.filter(school=school_admin.school)
    if request.method == 'POST':
        guardian = Guardian.objects.create(
            name=request.POST['name'],
            type=request.POST['type'],
            phone_number=request.POST['phone_number']
        )
        messages.success(request, 'Guardian Added Successfully')
        context = {
            'stream': Stream.objects.get(id=stream_id),
            'guardian': guardian,
            'fs': forms,
            'subjects': Subject.objects.all()
        }
        return render(request, 'school/school/add_student.html', context)
    context = {
        'stream': Stream.objects.get(id=stream_id),
    }
    return render(request, 'school/school/add_guardian.html', context)


@login_required()
def student_assign_form_stream(request, student_id):
    if request.method == 'POST':
        student_form = request.POST['student_form']
        year = request.POST['year']
        student_id = student_id
        stream = request.POST['stream']

        if StudentForm.objects.filter(student_id=student_id, year=year, is_current=True).count() > 0:
            StudentForm.objects.filter(student_id=student_id, year=year, is_current=True).update(form=student_form)
            if StudentStream.objects.filter(student_id=student_id).count() > 0:
                Student.objects.filter(student_id=student_id).update(stream=stream)
            else:
                Student.objects.create(
                    student_id=student_id,
                    stream=stream
                )
        else:
            student_forms = StudentForm.objects.filter(student_id=student_id)
            if student_forms is not None:
                for student_form in student_forms:
                    student_form.is_current = False
                StudentForm.objects.create(
                    student_id=student_id,
                    year=year,
                    is_current=True,
                )
            else:
                StudentForm.objects.create(
                    student_id=student_id,
                    year=year,
                    is_current=True,
                    form=student_form
                )
    pass


@login_required()
def add_term(request, form_id):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    if request.method == 'POST':
        name = request.POST['name']
        # year = request.POST['year']
        # fee = request.POST['fee']
        school_id = school.id
        term = Term.objects.create(
            name=name,
            # year=year,
            # fee=fee,
            school_id=school_id
        )
        term_year = TermYear.objects.create(
            date=datetime.now(),
            term=term
        )
        messages.success(request, 'Term Added Successfully')
        return redirect('School:exam_terms_by_year', term_year.date.year, form_id)
    context = {
        'school': school,
        'form_id': form_id
    }
    return render(request, 'school/school/add_term.html', context)


def school_term_years(request, term_id):
    term_years = TermYear.objects.filter(term_id=term_id)
    # years = []
    # for term_year in term_years:
    #     years.append(term_year.date.year)

    context = {
        'term_years': term_years
    }
    return render(request, 'school/finance/years.html', context)


def school_finance_classes(request, term_year_id):
    term_year = TermYear.objects.filter(id=term_year_id).first()
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    forms = Form.objects.filter(school=school)

    context = {
        'forms': forms,
        'term_year': term_year
    }
    return render(request, 'school/finance/school_classes.html', context)


def set_term_fee(request, term_year_id, form_id):
    s_term_year = TermYear.objects.filter(id=term_year_id).first()
    form = Form.objects.filter(id=form_id).first()
    s_fee = Fee.objects.filter(term_year=s_term_year).first()
    if request.method == 'POST':
        fee = request.POST['fee']
        form = request.POST['form']
        term_year = request.POST['term_year']

        if fee is not None:
            print(fee)
            s_fee.fee = fee
            s_fee.term_year_id = term_year
            s_fee.form_id = form
            s_fee.save()
        else:
            Fee.objects.create(
                fee=fee,
                form_id=form,
                term_year_id=term_year
            )
        context = {
            'term_year': s_term_year,
            'fee_break_downs': FeeBreakDown.objects.filter(term_year=term_year)
        }
        return render(request, 'school/finance/fee_break_down.html', context)

    context = {
        'term_year': s_term_year,
        'form': form,
        'fee': s_fee
    }
    return render(request, 'school/finance/set_term_fee.html', context)


def add_fee_break_down(request, term_year_id):
    term_year = TermYear.objects.filter(id=term_year_id).first()
    if request.method == 'POST':
        form = FeeForm(request.POST)
        if form.is_valid():
            form.save()

            context = {
                'term_year': term_year,
                'fee_break_downs': FeeBreakDown.objects.filter(term_year=term_year)
            }
            return render(request, 'school/finance/fee_break_down.html', context)
        else:
            print(form)
    context = {
        'term_year': term_year
    }
    return render(request, 'school/finance/add_fee_break_down.html', context)


@login_required()
def edit_term(request, term_id):
    term = Term.objects.get(id=term_id)
    if request.method == 'POST':
        term.name = request.POST['name']
        # term.year = request.POST['year']
        # term.fee = request.POST['fee']
        term.save()

    context = {
        'term': term
    }
    return render(request, '')


@login_required()
def examination_classes(request):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    classes = Form.objects.filter(school=school)

    context = {
        'classes': classes,
        'school': school
    }
    return render(request, 'school/exam/classes.html', context)


@login_required()
def exam_years(request):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)

    current_year = datetime.now().year
    start_year = school.created_at.year
    difference = current_year - start_year
    years = []
    years.append(current_year)

    while difference > 0:
        years.append(current_year - 1)
        difference -= 1

    context = {
        'years': years
    }
    return render(request, 'school/exam/years.html', context)


def forms_by_year(request, year):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    classes = Form.objects.filter(school=school)

    context = {
        'classes': classes,
        'school': school,
        'year': year
    }
    return render(request, 'school/exam/classes.html', context)


def exam_terms_by_year(request, year, form_id):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    terms_years = TermYear.objects.filter(term__school=school, date__year=year)
    term_ids = [term_year.term_id for term_year in terms_years]
    actual_terms = Term.objects.filter(id__in=term_ids)
    context = {
        'terms': actual_terms,
        'school': school,
        'form_id': form_id,
        'year': year
    }

    return render(request, 'school/exam/terms.html', context)


@login_required()
def examination_terms(request, form_id):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    terms = Term.objects.filter(school=school)

    context = {
        'terms': terms,
        'school': school,
        'form_id': form_id
    }

    return render(request, 'school/exam/terms.html', context)


@login_required()
def examinations_exam(request, term_year_id, form_id):
    term_year = TermYear.objects.filter(id=term_year_id).first()
    print(term_year)
    exams = Exam.objects.filter(term_year=term_year.id, form_id=form_id)
    context = {
        'exams': exams,
        'form_id': form_id,
        'term_year': term_year
    }

    return render(request, 'school/exam/exams.html', context)


@login_required()
def add_exam(request, form_id, term_year_id):
    term_year = TermYear.objects.filter(id=term_year_id).first()
    form = Form.objects.get(id=form_id)
    if request.method == 'POST':
        term_year_id = request.POST['term_year_id']
        name = request.POST['name']

        Exam.objects.create(
            term_year_id=term_year_id,
            name=name,
            form=form
        )
        messages.success(request, 'Exam Added Successfully')
        return redirect('School:examinations_exam', term_year.date.year, form_id)

    context = {
        'form': form,
        'term_year': term_year
    }

    return render(request, 'school/exam/add_exam.html', context)


@login_required()
def examination_students(request, form_id, exam_id):
    student_forms = StudentForm.objects.filter(form_id=form_id, is_current=True)
    ids = []
    for student_form in student_forms:
        ids.append(student_form.student.id)
    students = Student.objects.filter(id__in=ids)

    context = {
        'students': students,
        'exam_id': exam_id,
        'form_id': form_id
    }
    return render(request, 'school/exam/students.html', context)


@login_required()
def student_subject_scores(request, student_id, exam_id, form_id):
    student_subjects = StudentFormSubject.objects.filter(form_id=form_id, student_id=student_id)
    dummy_subjects = ExamStudentSubjectScore.objects.filter(student_id=student_id, exam_id=exam_id)
    subjects = []
    print(student_id, exam_id, form_id)
    for subject in student_subjects:
        print(subject.subject.id)
        if subject.subject.id in (x.subject_id for x in dummy_subjects):
            for d in dummy_subjects:
                if subject.subject.id == d.subject_id:
                    data = [subject.subject.id, subject.subject.name, d.score, d.grade]
                    subjects.append(data)
        else:
            data = [subject.subject.id, subject.subject.name, 0, 0]
            subjects.append(data)
    context = {
        'data': subjects,
        'student_id': student_id,
        'exam_id': exam_id
    }
    print(subjects)
    return render(request, 'school/exam/subject_scores.html', context)


@login_required()
def edit_subject_score(request, student_id, exam_id, subject_id):
    student_form = StudentForm.objects.get(student_id=student_id, is_current=True)
    if request.method == 'POST':
        score = request.POST['score']
        exam_subject_score = ExamStudentSubjectScore.objects.filter(student_id=student_id, exam_id=exam_id, subject_id=subject_id).first()
        if exam_subject_score is not None:
            exam_subject_score.score = score
            exam_subject_score.save()
        else:
            ExamStudentSubjectScore.objects.create(
                student_id=student_id,
                exam_id=exam_id,
                subject_id=subject_id,
                score=score
            )
        messages.success(request, 'Student Score Updated Successfully')
        return redirect('School:student_subject_scores', student_id, exam_id, student_form.form_id)

    context = {
        'student_id': student_id,
        'exam': Exam.objects.get(id=exam_id),
        'subject': Subject.objects.get(id=subject_id)
    }
    return render(request,'school/exam/edit_student_subject_score.html', context)


@login_required()
def generate_exam_results(request, exam_id):
    # file path
    exam = Exam.objects.filter(id=exam_id).first()
    time = datetime.datetime.now()
    dir_name = Path("media/exam_results/{}/{}".format(time.year, time.month))  # folder structure
    dir_name.mkdir(parents=True, exist_ok=True)
    time_only = "%d%d%d%d" % (time.day, time.hour, time.minute, time.second)
    filepath = "{}/{}_{}.xlsx".format(dir_name, exam.name.replace(' ', '_'), time_only)

    students = exam.sorted_students()

    # create a workbook
    workbook = Workbook()
    work_sheet = workbook.get_sheet_by_name('Sheet')
    work_sheet.title = 'Summary'

    work_sheet.append((' ', ' ', ''))
    work_sheet.append((' ', ' ', ''))
    work_sheet.append((' ', ' ', ''))
    work_sheet.append((' ', ' ', ''))
    work_sheet.append(('RANK', 'ADM NO', 'STUDENT NAME', 'SCORE', 'GRADE'))

    # print(students)
    data_rows = []
    for student in students:
        subject_scores = ExamStudentSubjectScore.objects.filter(exam_id=exam_id, student_id=student.id)
        final_score = 0
        individual_student = []
        individual_student.extend([student.score_rank(exam_id), student.admission_number, f'{student.first_name} {student.last_name}'])
        for score in subject_scores:
            final_score += score.score
        individual_student.extend([final_score,'A'])
        data_rows.append(individual_student)

    for row in data_rows:
        work_sheet.append(row)
    workbook.save(filepath)
    if os.path.exists(filepath):
        with open(filepath, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(filepath)
            return response
    return Http404


@login_required()
def delete_stream(request, stream_id):
    stream = Stream.objects.filter(id=stream_id).first()
    form_id = stream.form_id
    stream.delete()
    messages.success(request, 'Stream deleted successfully')
    return redirect('School:school_streams', form_id)


@login_required()
def delete_student(request, student_id):
    student = Student.objects.filter(id=student_id)
    student.delete()
    student_stream = StudentStream.objects.get(student_id=student_id)
    messages.success(request, 'Student deleted successfully')
    return redirect('School:school_students', student_stream.stream.id)


@login_required()
def delete_class(request, class_id):
    school_class = Form.objects.filter(id=class_id).first()
    school_class.delete()
    messages.success(request, 'Class deleted successfully')
    return redirect('School:school_classes')


def finance_years(request):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)

    current_year = datetime.now().year
    start_year = school.created_at.year
    difference = current_year - start_year
    years = [current_year]

    while difference > 0:
        years.append(current_year - 1)
        difference -= 1

    context = {
        'years': years
    }
    return render(request, 'school/finance/school_years.html', context)


@login_required()
def finance_year_classes(request, year):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    classes = Form.objects.filter(school=school_admin.school)

    context = {
        'classes': classes,
        'year': year
    }

    return render(request, 'school/finance/school_classes.html', context)


def finance_terms(request, form_id, year):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    school = School.objects.get(id=school_admin.school.id)
    terms_years = TermYear.objects.filter(term__school=school, date__year=year)
    term_ids = [term_year.term_id for term_year in terms_years]
    actual_terms = Term.objects.filter(id__in=term_ids)
    context = {
        'terms': actual_terms,
        'school': school,
        'form_id': form_id,
        'year': year
    }
    return render(request, 'school/finance/terms.html', context)


""" to be deleted"""
@login_required()
def finance_classes(request):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    classes = Form.objects.filter(school=school_admin.school)

    context = {
        'classes': classes
    }

    return render(request, 'school/finance/finance_classes.html', context)


@login_required()
def finance_students(request, class_id):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    class_students = StudentForm.objects.filter(form_id=class_id, form__school=school_admin.school, is_current=True)
    ids = []
    for c in class_students:
        ids.append(c.student_id)
    students = Student.objects.filter(id__in=ids)

    context = {
        'school': school_admin.school,
        'students': students,
        'class_id': class_id,
    }
    return render(request, 'school/finance/finance_students.html', context)


@login_required()
def finance_student_payments(request, student_id, term_id, form_id):
    payments = StudentFee.objects.filter(student_id=student_id, term_id=term_id, form_id=form_id)

    context = {
        'payments': payments,
        'student_id': student_id,
        'term_id': term_id,
        'form_id': form_id
    }
    return render(request, 'school/finance/finance_student_payments.html', context)


@login_required()
def finance_add_payment(request, student_id, term_id, form_id ):

    if request.method == 'POST':
        amount_paid = request.POST['amount_paid']
        date_paid = request.POST['date_paid']

        StudentFee.objects.update_or_create(
            student_id=student_id,
            term_id=term_id,
            form_id=form_id,
            amount_paid=amount_paid,
            date_paid=date_paid
        )
        messages.success(request, 'payment added successfully')
        return redirect('School:finance_student_payments', student_id, term_id, form_id)

    context = {
        'student_id': student_id,
        'term_id': term_id,
        'form_id': form_id
    }
    return render(request, 'school/finance/add_payment.html', context)


def sms_menu(request):
    return render(request, 'school/sms/sms_menu.html')


def personalized_sms_menu(request):
    return render(request, 'school/sms/personalized_sms_menu.html')


def simple_sms(request):
    return render(request, 'school/sms/simple_sms.html')


def personalized_from_file(request):
    return render(request, 'school/sms/personalized_from_file.html')


def personalized_from_contact_group(request):
    return render(request, 'school/sms/personalized_from_contact_group.html')


@login_required()
def import_csv_2(request):
    if request.method == 'POST' and request.FILES['myfile']:
        file = request.FILES['myfile']
        context = get_excel_content(file)
        return render(request, 'sms/sms.html', context)
    return render(request, 'sms/sms.html')


@login_required()
def merge_sms_2(request):
    data = []
    if request.method == 'POST':
        message = request.POST['Message']
        phone_number_field = request.POST['NumberField']
        file = request.POST['file_path']

        file_path = file.split('/', 1)[1]
        workbook = load_workbook('media/%s' % file_path)
        sheet_names = workbook.sheetnames

        sheet = sheet_names[0]
        worksheet = workbook.get_sheet_by_name(sheet)

        parameters = get_message_parameters(message=message)
        parameter_cells = get_parameter_column(parameters=parameters, worksheet=worksheet)
        phone_number_column = get_phone_number_column(phone_number_field=phone_number_field, worksheet=worksheet)

        max_row = worksheet.max_row
        max_column = worksheet.max_column
        for i in range(2, max_row + 1):
            person_message = {}
            new_message = message
            sms = ''
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)

                for a, b in parameter_cells.items():
                    if j == b:
                        new_message = new_message.replace('[%s]' % a, str(cell_obj.value))
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)
                if j == phone_number_column:
                    phone_number = cell_obj.value
                    sms = Merged(phone_number, new_message)
                    person_message['phone_number'] = phone_number
                    person_message['message'] = new_message
            data.append(sms)
    data_dict = {}
    for d in data:
        data_dict[d.phone_number] = d.message
    request.session['data'] = data_dict
    return render(request, 'sms/sample_merged_sms.html', {'data': data})


def grade_student(request, student_id, exam_id):
    points = 0
    # student = Student.objects.get(id=student_id)
    student_scores = ExamStudentSubjectScore.objects.filter(student_id=student_id, exam_id=exam_id)
    exam = Exam.objects.get(id=exam_id)
    # student_form_subject = StudentFormSubject.objects.filter(student_id=student_id, form_id=exam.form_id)
    final_score = 0
    group_two = 0
    group_three = 0
    group_four = 0
    used_scores = []
    for score in student_scores:
        if score.subject.subject_group.name == 'Compulsory':
            final_score += score.score
            points += get_points(score.score)
        elif score.subject.subject_group.name == 'Group 2':
            if group_two == 0:
                e = ExamStudentSubjectScore.objects.filter(student_id=student_id, exam_id=exam_id,
                                                           subject__subject_group__name='Group 2')
                scores = []
                for s in e:
                    scores.append(s.score)
                largest = Nmaxelements(scores, 2)
                # print('group2',largest)
                if largest:
                    for l in largest:
                        final_score += l
                        points += get_points(l)
                        for s in e:
                            if s.score == l:
                                used_scores.append(s.id)
                group_two = 1

        elif score.subject.subject_group.name == 'Group 3':
            if group_three == 0:
                e = ExamStudentSubjectScore.objects.filter(student_id=student_id, exam_id=exam_id,
                                                           subject__subject_group__name='Group 3')
                scores = []
                for s in e:
                    scores.append(s.score)

                largest = Nmaxelements(scores, 1)
                # print('group3',largest)
                if largest:
                    for l in largest:
                        final_score += l
                        points += get_points(l)
                        for s in e:
                            if s.score == l:
                                used_scores.append(s.id)
                group_three = 1
                # print('used_scores', used_scores)

        elif score.subject.subject_group.name == 'Group 2' or score.subject.subject_group.name == 'Group 3' or\
                score.subject.subject_group.name == 'Group 4' or score.subject.subject_group.name == 'Group 5':
            if group_four == 0 and group_two == 1 and group_three == 1:
                e = ExamStudentSubjectScore.objects.filter(Q(subject__subject_group__name='Group 2') |
                                                           Q(subject__subject_group__name='Group 3') |
                                                           Q(subject__subject_group__name='Group 4') |
                                                           Q(subject__subject_group__name='Group 5'),
                                                           student_id=student_id, exam_id=exam_id).exclude(id__in=used_scores)
                scores = []
                for s in e:
                    scores.append(s.score)
                largest = Nmaxelements(scores, 1)
                # print('group 4', largest)
                if largest:
                    for l in largest:
                        final_score += l
                        points += get_points(l)
                group_four = 1
    context = {
        'final_score': final_score,
        'points': points
    }
    return render(request, 'school/exam/final_score.html', context)


def get_points(score):
    grade = ''
    points = 0
    if 80 <= score <= 100:
        grade = 'A'
        points = 12
    elif 75 <= score <= 79.99:
        grade = 'A-'
        points = 11
    elif 70 <= score <= 74.99:
        grade = 'B+'
        points = 10
    elif 65 <= score <= 69.99:
        grade = 'B'
        points = 9
    elif 60 <= score <= 64.99:
        grade = 'B-'
        points = 8
    elif 55 <= score <= 59.99:
        grade = 'C+'
        points = 7
    elif 50 <= score <= 54.99:
        grade = 'C'
        points = 6
    elif 45 <= score <= 49.99:
        grade = 'C-'
        points = 5
    elif 40 <= score <= 44.99:
        grade = 'D+'
        points = 4
    elif 35 <= score <= 39.99:
        grade = 'D'
        points = 3
    elif 30 <= score <= 34.99:
        grade = 'D-'
        points = 2
    elif 0 <= score <= 29.99:
        grade = 'E'
        points = 1
    return points


def per_subject_grading(score):
    grade = ''
    points = 0
    if 80 <= score <= 100:
        grade = 'A'
        points = 12
    elif 75 <= score <= 79.99:
        grade = 'A-'
        points = 11
    elif 70 <= score <= 74.99:
        grade = 'B+'
        points = 10
    elif 65 <= score <= 69.99:
        grade = 'B'
        points = 9
    elif 60 <= score <= 64.99:
        grade = 'B-'
        points = 8
    elif 55 <= score <= 59.99:
        grade = 'C+'
        points = 7
    elif 50 <= score <= 54.99:
        grade = 'C'
        points = 6
    elif 45 <= score <= 49.99:
        grade = 'C-'
        points = 5
    elif 40 <= score <= 44.99:
        grade = 'D+'
        points = 4
    elif 35 <= score <= 39.99:
        grade = 'D'
        points = 3
    elif 30 <= score <= 34.99:
        grade = 'D-'
        points = 2
    elif 0 <= score <= 29.99:
        grade = 'E'
        points = 1
    grading = {}
    grading['grade'] = grade
    grading['points'] = points
    return grading


# Function returns N largest elements
def Nmaxelements(list1, N):
    final_list = []

    for i in range(0, N):
        max1 = 0
        for j in range(len(list1)):
            if list1[j] > max1:
                max1 = list1[j]

        list1.remove(max1)
        final_list.append(max1)
    return final_list
    # print(final_list)


def download_exam_template(request, exam_id):
    exam = Exam.objects.filter(id=exam_id).first()
    student_forms = StudentForm.objects.filter(form_id=exam.form_id)
    students = []
    for student_form in student_forms:
        students.append(student_form.student)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Exam Results'
    titles = ['ADM', 'STUDENT NAME']
    for subject in Subject.objects.all():
        titles.append(subject.name)
    worksheet.append(titles)
    for student in list(set(students)):
        worksheet.append((
            student.admission_number,
            f"{student.first_name} {student.last_name}",
        ))

    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/ms-excel')

    response['Content-Disposition'] = 'attachment; filename=template.xlsx'
    return response


@login_required()
def upload_student_marks(request, exam_id):
    exam = Exam.objects.filter(id=exam_id).first()
    student_forms = StudentForm.objects.filter(form_id=exam.form_id)
    students = []
    for student_form in student_forms:
        students.append(student_form.student)

    if request.method == 'POST':
        file = request.FILES['my_file']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        uploaded_file_url = fs.url(filename)
        file_path = uploaded_file_url.split('/', 1)[1]

        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
        target_sheet = sheet_names[0]
        worksheet = workbook[target_sheet]
        first_row = True
        for i in range(1, worksheet.max_row + 1):
            # for k in range(2, worksheet.max_column):
            if first_row:
                first_row = False
                continue
            else:
                print('row 2')
                first_column = True
                second_column = True
                student_id = 0
                for k in range(1, worksheet.max_column):

                    if first_column:
                        first_column = False
                        student_id = Student.objects.filter(
                            admission_number=worksheet.cell(row=i, column=k).value).first().id
                        continue
                    elif second_column:
                        second_column = False
                        continue
                    if not first_column and not second_column:
                        print('column 3')
                        ExamStudentSubjectScore.objects.update_or_create(
                            exam=exam,
                            subject=Subject.objects.filter(name=worksheet.cell(row=1, column=k).value).first(),
                            score=worksheet.cell(row=i, column=k).value,
                            defaults={
                                "student_id": student_id,
                            }
                        )
                        print('got here')
        messages.success(request, 'Upload Complete')
        return redirect('School:examinations_exam', exam.form_id, exam.term_id)


def payment_receipt_preview(request, student_fee_id):
    payment = StudentFee.objects.filter(id=student_fee_id).first()
    return render(request, 'school/finance/payment_receipt.html', {'payment':payment})


def settings(request):
    school_admin = SchoolAdmin.objects.get(user_ptr_id=request.user.id)
    classes = Form.objects.filter(school=school_admin.school)
    terms = Term.objects.filter(school=school_admin.school)
    context = {
        'classes': classes,
        'terms': terms,
    }
    print(terms)
    return render(request, 'school/settings/settings.html', context)


def term_years(request, term_id):
    term = Term.objects.filter(id=term_id).first()
    term_years = TermYear.objects.filter(term_id=term_id)

    actual_years = []
    for term_year in term_years:
        actual_years.append(term_year.date.year)
    context = {
        'years': actual_years
    }
    return render(request, 'school/settings/term_years.html', context)


    # extension = file.name.rsplit('.', 1)[1]
    # if extension == 'csv':
    #     file_path = uploaded_file_url.split('/', 1)[1]
    #     with open(file_path, 'r') as f:
    #         firstline = True
    #         for row in csv.reader(f):
    #             if firstline:
    #                 firstline = False
    #                 continue
    #             else:
    #                 p = f"{254}{row[1].replace(' ', '')[-9:]}"
    #                 Contact.objects.update_or_create(
    #                     name=row[0],
    #                     group_id=group_id,
    #                     phone_number=int(p),
    #                     email=row[2]
    #                 )
    # else:
