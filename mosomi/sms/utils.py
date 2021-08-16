import csv
import datetime
import hashlib
import json
from pathlib import Path

import requests
from celery.task import task, periodic_task
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook
from pip._vendor.pyparsing import pyparsing_common
from zeep import Client

from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template, render_to_string
from xhtml2pdf import pisa
from collections import defaultdict
from django.apps import apps

from sms.models import OutgoingDone


def subtract_months(time, months):
    days = months * 30
    return time - datetime.timedelta(days=days)


def subtract_weeks(time, weeks):
    days = weeks * 7
    return time - datetime.timedelta(days=days)


def get_last_n_weeks(n):
    weeks = []
    i = 0
    while i < n + 1:
        date = subtract_weeks(datetime.datetime.today(), i)
        weeks.append(date.isocalendar()[1])
        i += 1
    return weeks


def get_last_n_months(n):
    months = []
    i = 0
    today = datetime.datetime.today()
    months.append(today.month)
    while i < n + 1:
        date = subtract_months(datetime.datetime.today(), i)
        months.append(date.month)
        i += 1
    return months


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


def write_pdf(template_src, context_dict, file_path, file_name):
    Path(file_path).mkdir(parents=True, exist_ok=True)
    output_filename = "%s/%s" % (file_path, file_name)
    result_file = open(output_filename, "w+b")
    pisa_status = pisa.CreatePDF(
        render_to_string(template_src, context_dict), dest=result_file
    )
    result_file.close()
    return output_filename


# def get_last_tem_weeks():
#     today = datetime.


def calculate_message_cost(message):
    '''
    Method to calculate text message cost
    Sample Usage:
        message = "Hello Simon"
        message_cost = calculate_message_cost(message)
        print(message_cost)

        output:
            1
    '''
    message_length = len(message)
    cost = 0

    if message_length >= 160:
        cost += 1
        new_length = message_length - 160
        if new_length >= 144:
            cost += 1
            new_length = new_length - 144
            if new_length >= 151:
                while new_length >= 151:
                    new_length = new_length - 151
                    cost += 1
                else:
                    cost += 1
            else:
                cost += 1
        else:
            cost += 1
    else:
        cost = 1
    return cost


def get_message_parameters(message):
    '''
    This method returns all parameters passed in the message specified by a user
    sample usage
        parameters = get_message_parameters(message)
        eg:
        message = Hello [name], check your email [email] for more information
        print(get_message_parameters(message))

        output:
        ['name', 'email']
    '''
    string_list = message.split()
    parameters = []
    for string in string_list:
        correct_parameter = ''
        if string.endswith('.'):
            without_full_stop = string.strip().replace('.', '')
            if without_full_stop.endswith(']') and without_full_stop.startswith('['):
                l = string.replace('[', '')
                k = l.replace(']', '')
                correct_parameter = k

        elif string.endswith(','):
            without_comma = string.replace(',', '')
            if without_comma.endswith(']') and without_comma.startswith('['):
                l = string.replace('[', '')
                k = l.replace(']', '')
                correct_parameter = k
            # print(parameters)
        else:
            if string.endswith(']') and string.startswith('['):
                l = string.replace('[', '')
                k = l.replace(']', '')
                correct_parameter = k
            # print(parameters)
        if correct_parameter != '':
            if correct_parameter.endswith('.'):
                my_parameter = correct_parameter.replace('.', '')
                parameters.append(my_parameter)
            elif correct_parameter.endswith(','):
                my_parameter = correct_parameter.replace(',', '')
                parameters.append(my_parameter)
            else:
                my_parameter = correct_parameter
                parameters.append(my_parameter)
    return parameters


def get_phone_number_column(phone_number_field, worksheet):
    '''
    This method returns the column corresponding to the phone number field specified by the user(client)
    '''
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    phone_number_column = ''
    for i in range(1, 2):
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            if cell_obj.value == phone_number_field:
                phone_number_column = j
    if phone_number_column != '':
        return phone_number_column


def get_parameter_column(parameters, worksheet):
    '''
    For each parameter returned on the get_message_parameters() method
    return the corresponding column in the uploaded excel file
    '''
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    parameter_cells = {}

    for parameter in parameters:
        for i in range(1, 2):
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)
                if cell_obj.value is not None:
                    if cell_obj.value.strip() == parameter.strip():
                        parameter_cells[parameter] = j
    return parameter_cells


def get_excel_content(file):
    """
    Save, Read an xlsx file and return its contents row by row
    """
    fs = FileSystemStorage()
    filename = fs.save(file.name, file)
    uploaded_file_url = fs.url(filename)

    extension = file.name.rsplit('.', 1)[1]
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        return convert_csv_to_xlsx(file_path)

    file_path = uploaded_file_url.split('/', 1)[1]
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    sheet = sheet_names[0]
    worksheet = workbook.get_sheet_by_name(sheet)

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    contacts = []

    for i in range(1, max_row + 1):
        person_details = []
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            person_details.append(cell_obj.value)
        contacts.append(person_details)

    headers = []
    for i in range(1, 2):
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            headers.append(cell_obj.value)
    context = {
        'contacts': contacts,
        'fields': headers,
        'file': file_path
    }

    return context


# convert csv to xlsx
def convert_csv_to_xlsx(file):
    """
    Method to convert uploaded .csv file to .xlsx before reading with openpyxl
     sample usage:
        file = convert_csv_to_xlsx(file_path)
    """
    f = file.rsplit('/', 1)[1]
    file_name = f.rsplit('.', 1)[0]
    wb = Workbook()
    ws = wb.active
    with open(file, 'r') as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save('media/%s.xlsx' % file_name)

    uploaded_file_url = 'media/%s.xlsx' % file_name
    return get_excel_content_after_conversion(uploaded_file_url)


def get_excel_content_after_conversion(file_path):
    '''
    Save, Read an xlsx file and return its contents row by row
    '''
    workbook = load_workbook(file_path)

    sheet_names = workbook.sheetnames
    sheet = sheet_names[0]
    worksheet = workbook.get_sheet_by_name(sheet)

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    contacts = []

    for i in range(1, max_row + 1):
        person_details = []
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            person_details.append(cell_obj.value)
        contacts.append(person_details)

    headers = []
    for i in range(1, 2):
        for j in range(1, max_column + 1):
            cell_obj = worksheet.cell(row=i, column=j)
            headers.append(cell_obj.value)
    context = {
        'contacts': contacts,
        'fields': headers,
        'file': file_path
    }
    return context


def convert(lst):
    return eval(lst)


def get_access_token():
    client_id = "dHq5I3HZDyR5vuh2wNhjGjSIBVCdNeie"
    client_secret = "91B9moegcTJSCmtlOw4YmohnxXpJ6doZdxryVY12"
    url = "https://api.emalify.com/v1/oauth/token"
    headers = {
        'Accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials"
    }
    response = requests.post(url, json=data, headers=headers)
    print(response.text)
    return json.loads(response.text)['access_token']


@csrf_exempt
class SDP:
    '''
    Class to generate send sms and get sms delivery status soap requests
    '''
    SP_ID = '601515'
    SP_PASSWORD = 'Vfrcdexsw12#'

    def get_sms_delivery_status(self, service_id, phone_number, request_identifier):
        timestamp = datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')
        m = hashlib.md5()
        combined_string = self.SP_ID + self.SP_PASSWORD + timestamp
        m.update(combined_string.encode('utf-8'))
        sp_password = m.hexdigest()

        bodyxml = '<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" ' \
                  'xmlns:v2="http://www.huawei.com.cn/schema/common/v2_1" ' \
                  'xmlns:loc="http://www.csapi.org/schema/parlayx/sms/send/v2_2/local">' \
                  '<soapenv:Header>' \
                  '<v2:RequestSOAPHeader>' \
                  f'<v2:spId>{self.SP_ID}</v2:spId>' \
                  f'<v2:spPassword>{sp_password}</v2:spPassword>' \
                  f'<v2:serviceId>{service_id}</v2:serviceId>' \
                  f'<v2:timeStamp>{timestamp}</v2:timeStamp>'

        bodyxml += f'<v2:OA>{phone_number}</v2:OA><v2:FA>{phone_number}</v2:FA>'
        bodyxml += '</v2:RequestSOAPHeader>' \
                   '</soapenv:Header>' \
                   '<soapenv:Body>' \
                   '<loc:getSmsDeliveryStatus>' \
                   f'<loc:requestIdentifier>{request_identifier}</loc:requestIdentifier>' \
                   '</loc:getSmsDeliveryStatus>' \
                   '</soapenv:Body>' \
                   '</soapenv:Envelope>'

        headers = {'Content-type': 'text/xml; charset=utf-8'}
        response = requests.post(url='http://41.90.0.130:8310/SendSmsService/services/SendSms', data=bodyxml, headers=headers)
        return response

    def send_sms_customized(self, service_id, recipients, message, sender_code):
        if not sender_code:
            sender_code = '711037'
        timestamp = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d%H%M%S')
        m = hashlib.md5()
        combined_string = self.SP_ID + self.SP_PASSWORD + timestamp
        m.update(combined_string.encode('utf-8'))
        sp_password = m.hexdigest()

        # soap request packaging
        body_xml = \
            "<soapenv:Envelope xmlns:soapenv='http://schemas.xmlsoap.org/soap/envelope/' " \
            "xmlns:v2='http://www.huawei.com.cn/schema/common/v2_1' " \
            "xmlns:loc='http://www.csapi.org/schema/parlayx/sms/send/v2_2/local'> " \
            "<soapenv:Header>"\
            "<v2:RequestSOAPHeader>"\
            f"<v2:spId>{self.SP_ID}</v2:spId>"\
            f"<v2:spPassword>{sp_password}</v2:spPassword>"\
            f"<v2:serviceId>{service_id}</v2:serviceId>"\
            f"<v2:timeStamp>{timestamp}</v2:timeStamp>"
        if len(recipients) == 1:
            body_xml += f"<v2:OA>tel:{recipients[0]}</v2:OA><v2:FA>tel:{recipients[0]}</v2:FA>"
        body_xml += "</v2:RequestSOAPHeader></soapenv:Header><soapenv:Body><loc:sendSms>"
        if len(recipients) == 1:
            body_xml += f"<loc:addresses>tel:{recipients[0]}</loc:addresses>"
        else:
            for recipient in recipients:
                body_xml += f"<loc:addresses>tel:{recipient}</loc:addresses>"
        body_xml += f"<loc:senderName>{sender_code}</loc:senderName><loc:message>{message}</loc:message>"
        body_xml += "</loc:sendSms></soapenv:Body></soapenv:Envelope>"
        headers = {'Content-type': 'text/xml; charset=utf-8'}
        response = requests.post(url='http://41.90.0.130:8310/SendSmsService/services/SendSms',
                                 data=body_xml,
                                 headers=headers)
        return response


class BulkCreateManager(object):
    """
    This helper class keeps track of ORM objects to be created for multiple
    model classes, and automatically creates those objects with `bulk_create`
    when the number of objects accumulated for a given model class exceeds
    `chunk_size`.
    Upon completion of the loop that's `add()`ing objects, the developer must
    call `done()` to ensure the final set of objects is created for all models.
    """

    def __init__(self, chunk_size=100):
        self._create_queues = defaultdict(list)
        self.chunk_size = chunk_size

    def _commit(self, model_class):
        model_key = model_class._meta.label
        model_class.objects.bulk_create(self._create_queues[model_key])
        counter = 1
        print(f'commit{counter}')
        counter += 1
        self._create_queues[model_key] = []

    def add(self, obj):
        """
        Add an object to the queue to be created, and call bulk_create if we
        have enough objs.
        """
        model_class = type(obj)
        model_key = model_class._meta.label
        self._create_queues[model_key].append(obj)
        if len(self._create_queues[model_key]) >= self.chunk_size:
            self._commit(model_class)

    def done(self):
        """
        Always call this upon completion to make sure the final partial chunk
        is saved.
        """
        for model_name, objs in self._create_queues.items():
            if len(objs) > 0:
                self._commit(apps.get_model(model_name))