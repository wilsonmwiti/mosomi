import json
import os
import random
from pprint import pprint
import datetime

from celery.result import AsyncResult
from celery.task import task
from django.contrib import messages
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.sites.shortcuts import get_current_site
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db import connections
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import CreateView, ListView
from django_chunked_iterator import iterator
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from rest_framework.response import Response

from mpesa_api.models import Mpesa
from roberms.celery import app
from sms.forms import *
from sms.models import *
from sms.tokens import account_activation_token
from sms.utils import *
from sms_api.views import send_usage
#from cgi import escape
import html
from html import escape

#html.escape(str).encode('ascii', 'xmlcharrefreplace')



def timed(f):
    @wraps(f)
    def wrapper(*args, **kwds):
        start = time()
        result = f(*args, **kwds)
        elapsed = time() - start
        print("%s took %d time to finish" % (f.__name__, elapsed))
        return result
    return wrapper


def is_user_customer(function):
    def wrap(request, *args, **kwargs):
        user = request.user
        customer = Customer.objects.filter(user_ptr_id=user.id).first()
        if customer is not None:
            return function(request, *args, **kwargs)
        elif CustomerSubAccounts.objects.filter(user_ptr_id=user.id).first():
            return function(request, *args, **kwargs)
        else:
            return redirect('sms:login')
    return wrap


class Merged:
    def __init__(self, phone_number, message):
        self.phone_number = phone_number
        self.message = message


# @login_required()
def home(request):
    return render(request, 'sms/sample_home.html')


def about(request):
    return render(request, 'sms/about.html')


@login_required()
@is_user_customer
def apps(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is None:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        weeks = get_last_n_weeks(12)
        one_week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
        current_day = datetime.datetime.today()
        credit_usage = []
        for week in weeks:
            # print(week)
            messages = OutgoingDone.objects.filter(sent_time__gte=one_week_ago, sent_time__lte=current_day,
                                                   customer_id=customer.id).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_week_ago = one_week_ago - datetime.timedelta(days=7)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count()+1
        }
        return render(request, 'sms/apps.html', context)
    else:
        weeks = get_last_n_weeks(12)
        one_week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
        current_day = datetime.datetime.today()
        credit_usage = []
        for week in weeks:
            # print(week)
            messages = OutgoingDone.objects.filter(sent_time__gte=one_week_ago, sent_time__lte=current_day,
                                                   customer_id=customer.id).count()
            credit_usage.append(messages)
            current_day = current_day - datetime.timedelta(days=7)
            one_week_ago = one_week_ago - datetime.timedelta(days=7)
        # print(credit_usage)
        context = {
            'messages_sent': credit_usage[::-1],
            'weeks': weeks[::-1],
            'customer': customer,
            'contacts': Contact.objects.filter(group__customer_id=customer.id).count(),
            'groups': Group.objects.filter(customer_id=customer.id).count(),
            'admins': CustomerSubAccounts.objects.filter(owner=customer.id).count() + 1
        }
        return render(request, 'sms/apps.html', context)


@login_required()
@is_user_customer
def monthly_messages_sent(request):
    return


@login_required()
@is_user_customer
def services(request):
    return render(request, 'sms/services.html')


@login_required()
@is_user_customer
def contact(request):
    return render(request, 'sms/contact.html')


# @login_required()
class SendSmsView(LoginRequiredMixin, CreateView):
    model = Outgoing
    fields = ['phone_numbers', 'text_message']

    def form_valid(self, form):
        m = form.cleaned_data['phone_numbers'].splitlines()

        # for n in form.cleaned_data['phone_numbers'].splitlines():
        for n in m:
            form.instance.phone_numbers=n
            form.instance.text_message=form.cleaned_data['text_message']
            form.instance.user = self.request.user
            form.instance.access_code=self.request.user.profile.access_code
            form.instance.service_id=self.request.user.profile.service_id
        return super().form_valid(form)


# @login_required()
class SmsReportView(ListView):
    model = Outgoing

from functools import wraps
from time import time



@timed
@login_required()
@is_user_customer
def send(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        credit = customer.credit
        if request.method == 'POST':
            data = request.session['simple_messages']
            total_message_cost = 0
            for a, b in data.items():
                message_cost = calculate_message_cost(b)
                total_message_cost += message_cost

            if customer.credit >= total_message_cost:
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)
                track_code = trackingcode

                simple_sms_store.delay(customer.id, total_message_cost, data, track_code)

                request.session.delete("simple_messages")
                return redirect('sms:customer_reports', track_code)
            else:
                messages.error(request, 'You do not have enough credit in your account to make this request. '
                                        'Please Recharge To Continue')
                context = {
                    'phone_numbers': request.POST.get('phone_numbers'),
                    'message': request.POST.get('text_message'),
                    'customer': customer
                }
                return render(request, 'sms/simple_message.html', context)
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_message.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        credit = customer.credit
        if request.method == 'POST':
            data = request.session['simple_messages']
            total_message_cost = 0
            for a, b in data.items():
                message_cost = calculate_message_cost(b)
                total_message_cost += message_cost

            if customer.credit >= total_message_cost:
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)
                track_code = trackingcode

                simple_sms_store.delay(customer.id, total_message_cost, data, track_code)

                request.session.delete("simple_messages")
                return redirect('sms:customer_reports', track_code)
            else:
                messages.error(request, 'You do not have enough credit in your account to make this request. '
                                        'Please Recharge To Continue')
                context = {
                    'phone_numbers': request.POST.get('phone_numbers'),
                    'message': request.POST.get('text_message'),
                    'customer': customer
                }
                return render(request, 'sms/simple_message.html', context)
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_message.html', context)


@task
def simple_sms_store(customer_id, total_message_cost, data, track_code):
    customer = Customer.objects.filter(id=customer_id).first()
    new_credit = customer.credit - total_message_cost
    customer.credit = new_credit
    customer.save()

    bulk_mgr = BulkCreateManager(chunk_size=100)
    for a, b in data.items():
        s = ''.join(a.split())
        p = f"{254}{s[-9:]}"

        # if str.isdigit(customer.access_code):
        #     bulk_mgr.add(Outgoing(
        #         customer=customer,
        #         service_id=customer.service_id,
        #         access_code=customer.access_code,
        #         phone_number=p,
        #         text_message=b,
        #         track_code=track_code
        #     ))
        # else:
        bulk_mgr.add(OutgoingNew(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=p,
            text_message=b,
            track_code=track_code,
            sent_time=timezone.now()
        ))
    bulk_mgr.done()
    send_usage(total_message_cost)
    return 'insertion complete'


@login_required()
@is_user_customer
def simple_sms(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        if request.method == 'POST':
            message = request.POST.get('text_message')
            phone_numbers = request.POST.get('phone_numbers').splitlines()
            data = {}
            # print(phone_numbers)
            new_phone_numbers = []
            for phone_number in phone_numbers:
                if phone_number != '':
                    new_phone_numbers.append(phone_number)
            message_cost = calculate_message_cost(message)
            total_message_cost = message_cost * len(new_phone_numbers)

            if customer.credit >= total_message_cost:
                for p in new_phone_numbers:
                    # pprint(p)
                    data[p] = message
                # pprint(data)
                request.session['simple_messages'] = data
                return redirect("sms:simple_sms_preview")
            else:
                messages.error(request, 'You do not have enough credit to perform this action. Kindly Top Up To Continue')
                return redirect("sms:simple_sms")
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_sms.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        if request.method == 'POST':
            message = request.POST.get('text_message')
            phone_numbers = request.POST.get('phone_numbers').splitlines()
            data = {}
            # print(phone_numbers)
            new_phone_numbers = []
            for phone_number in phone_numbers:
                if phone_number != '':
                    new_phone_numbers.append(phone_number)
            message_cost = calculate_message_cost(message)
            total_message_cost = message_cost * len(new_phone_numbers)

            if customer.credit >= total_message_cost:
                for p in new_phone_numbers:
                    # pprint(p)
                    data[p] = message
                # pprint(data)
                request.session['simple_messages'] = data
                return redirect("sms:simple_sms_preview")
            else:
                messages.error(request,
                               'You do not have enough credit to perform this action. Kindly Top Up To Continue')
                return redirect("sms:simple_sms")
        else:
            context = {
                'customer': customer
            }
            return render(request, 'sms/simple_sms.html', context)


@login_required()
@is_user_customer
def simple_sms_preview(request):
    # print(request.session['simple_messages'])
    context = {
        'data': request.session['simple_messages']
    }
    return render(request, 'sms/simple_sms_preview.html', context)


@login_required()
@is_user_customer
def import_csv_2(request):
    if request.method == 'POST' and request.FILES['myfile']:
        file = request.FILES['myfile']
        context = get_excel_content(file)
        return render(request, 'sms/sms.html', context)
    return render(request, 'sms/sms.html')


@login_required()
@is_user_customer
def messages_dashboard(request):
    return render(request, 'sms/dashboard.html')


@login_required()
@is_user_customer
def merge_sms_2(request):
    data = []
    if request.method == 'POST':
        message = request.POST['Message']
        phone_number_field = request.POST['NumberField']
        file = request.POST['file_path']

        file_path = file.split('/', 1)[1]
        workbook = load_workbook('media/%s' % file_path)
        sheet_names = workbook.sheetnames

        sheet = sheet_names[0]
        worksheet = workbook.get_sheet_by_name(sheet)

        parameters = get_message_parameters(message=message)
        parameter_cells = get_parameter_column(parameters=parameters, worksheet=worksheet)
        phone_number_column = get_phone_number_column(phone_number_field=phone_number_field, worksheet=worksheet)

        max_row = worksheet.max_row
        max_column = worksheet.max_column
        for i in range(2, max_row + 1):
            person_message = {}
            new_message = message
            sms = ''
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)

                for a, b in parameter_cells.items():
                    if j == b:
                        new_message = new_message.replace('[%s]' % a, str(cell_obj.value))
            for j in range(1, max_column + 1):
                cell_obj = worksheet.cell(row=i, column=j)
                if j == phone_number_column:
                    phone_number = cell_obj.value
                    sms = Merged(phone_number, new_message)
                    person_message['phone_number'] = phone_number
                    person_message['message'] = new_message
            data.append(sms)
    data_dict = {}
    for d in data:
        data_dict[d.phone_number] = d.message
    request.session['data'] = data_dict
    return render(request, 'sms/sample_merged_sms.html', {'data': data})


@login_required()
@is_user_customer
def confirm(request):
    # pprint(request.session['data'])
    if request.method == 'POST':
        tracking_code = random.randint(1, 1000000)
        while OutgoingNew.objects.filter(track_code=tracking_code).count() > 0:
            tracking_code = random.randint(1, 1000000)
        phone_numbers = request.POST.getlist('phone_numbers[]')
        c_messages = request.session['data']
        # pprint(c_messages)
        customer = Customer.objects.filter(user_ptr_id=request.user).first()
        if customer is not None:
            actual_messages_cost = 0
            for a, message in c_messages.items():
                message_cost = calculate_message_cost(message)
                actual_messages_cost += message_cost

            if customer.credit >= actual_messages_cost:
                bulk_mgr = BulkCreateManager(chunk_size=100)
                for a, message in c_messages.items():
                    p = f"{254}{a.replace(' ', '')[-9:]}"

                    # if str.isdigit(customer.access_code):
                    #     bulk_mgr.add(Outgoing(
                    #         customer=customer,
                    #         service_id=customer.service_id,
                    #         access_code=customer.access_code,
                    #         phone_number=p,
                    #         text_message=message,
                    #         track_code=tracking_code
                    #     ))
                    # else:
                    bulk_mgr.add(OutgoingNew(
                        customer=customer,
                        service_id=customer.service_id,
                        access_code=customer.access_code,
                        phone_number=p,
                        text_message=message,
                        track_code=tracking_code,
                        sent_time=timezone.now()
                    ))

                    # bulk_mgr.add(Outgoing(phone_number=p,
                    #                       text_message=message,
                    #                       service_id=customer.service_id,
                    #                       access_code=customer.access_code,
                    #                       customer_id=customer.id,
                    #                       track_code=tracking_code))
                bulk_mgr.done()
                customer.credit = customer.credit - actual_messages_cost
                customer.save()
                send_usage(actual_messages_cost)
                return redirect('sms:customer_reports', tracking_code)
            else:
                data = []
                data_dict = request.session.get('data')
                for a, b in data_dict.items():
                    sms = Merged(a, b)
                    data.append(sms)
                messages.warning(request,
                                 'You do not have enough credit in your account to perform this action please recharge to continue')
                return render(request, 'sms/sample_merged_sms.html', {'data': data})
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user).first().owner
            actual_messages_cost = 0
            for a, message in c_messages.items():
                message_cost = calculate_message_cost(message)
                actual_messages_cost += message_cost

            if customer.credit >= actual_messages_cost:
                bulk_mgr = BulkCreateManager(chunk_size=100)
                for a, message in c_messages.items():
                    p = f"{254}{a.replace(' ', '')[-9:]}"
                    # bulk_mgr.add(Outgoing(phone_number=p,
                    #                       text_message=message,
                    #                       service_id=customer.service_id,
                    #                       access_code=customer.access_code,
                    #                       customer_id=customer.id,
                    #                       track_code=tracking_code))
                    #
                    # if str.isdigit(customer.access_code):
                    #     bulk_mgr.add(Outgoing(
                    #         customer=customer,
                    #         service_id=customer.service_id,
                    #         access_code=customer.access_code,
                    #         phone_number=p,
                    #         text_message=message,
                    #         track_code=tracking_code
                    #     ))
                    # else:
                    bulk_mgr.add(OutgoingNew(
                        customer=customer,
                        service_id=customer.service_id,
                        access_code=customer.access_code,
                        phone_number=p,
                        text_message=message,
                        track_code=tracking_code,
                        sent_time=timezone.now()
                    ))
                bulk_mgr.done()
                customer.credit = customer.credit - actual_messages_cost
                customer.save()
                send_usage(actual_messages_cost)
                return redirect('sms:customer_reports', tracking_code)
            else:
                data = []
                data_dict = request.session.get('data')
                for a, b in data_dict.items():
                    sms = Merged(a, b)
                    data.append(sms)
                messages.warning(request,
                                 'You do not have enough credit in your account to perform this action please recharge to continue')
                return render(request, 'sms/sample_merged_sms.html', {'data': data})
    return render(request, 'sms/result.html')


def register(request):
    if request.method == "POST":
        sdp = SDP()
        customer_code = random.randint(10000, 99999)
        while Customer.objects.filter(customer_code=customer_code).count() > 0:
            customer_code = random.randint(10000, 99999)
        form = CustomerRegisrationForm(request.POST)
        if form.is_valid():
            if Customer.objects.filter(username=request.POST['email']).count() < 1:
                customer = form.save(commit=False)
                customer.first_name = request.POST['username']
                customer.username = request.POST['email']
                customer.email = request.POST['email']
                customer.is_active = False
                customer.customer_code = customer_code
                customer.access_code = 'ROBERMS_LTD'
                customer.sender_name = 'ROBERMS_LTD'
                customer.save()
                phone_number = f"{254}{customer.phone_number.replace(' ', '')[-9:]}"
                OutgoingNew.objects.create(phone_number=phone_number,
                                      text_message="Account created successfully!\n" +
                                                   f"Verify the account with this code {customer.customer_code} now.\n" +
                                                   f"If you are verifying later, use the link below to activate \n" +
                                                   f"https://roberms.co.ke/verify/account",
                                      service_id=6015152000175328,
                                      access_code='ROBERMS_LTD',
                                      customer_id=customer.id,
                                      track_code=customer_code,
                                           sent_time=timezone.now())
                # response = sdp.send_sms_customized(service_id='6015152000175328', recipients=[phone_number],
                #                                	message=f'Verfication Code {customer.customer_code}', sender_code='711037')
                messages.success(request, 'Your account has been created! We sent you an text message containing a verification code')
                return redirect('sms:verify_account')
            else:
                messages.error(request, 'That Email Is Already Registered To Our System')
                return render(request, 'registration/register.html')
        else:
            # print('me')
            # print(form.errors)
            return render(request, 'registration/register.html', {'form': form})
    else:
        form = CustomerRegisrationForm()
        return render(request, 'registration/register.html', {'form': form})


def verify_account(request):
    if request.method == 'POST':
        if Customer.objects.filter(customer_code=request.POST['customer_code'], is_active=False).count() > 0:
            customer = Customer.objects.filter(customer_code=request.POST['customer_code']).first()
            customer.is_active = True
            customer.save()
            messages.success(request, f'Welcome {customer.first_name} your account is now active')
            return redirect('sms:login')
        else:
            messages.error(request, 'That code verification is not valid, please try again ')
            return render(request, 'registration/verify_account.html')
    return render(request, 'registration/verify_account.html')


@login_required()
@is_user_customer
def profile(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    return render(request, 'sms/profile.html', {'customer': customer})


@login_required()
@is_user_customer
def edit_profile(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if request.method == 'POST':
        email = request.POST['email']
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        business_name = request.POST['business_name']
        customer.email = email
        customer.first_name = first_name
        customer.last_name = last_name
        customer.business_name = business_name
        customer.save()
        messages.success(request, 'Profile Updated Successfully')
        return redirect('sms:apps')
    return render(request, 'sms/edit_profile.html', {'customer': customer})


def customer_credit(request):

    return render(request, 'sms/credit.html')


@login_required()
@is_user_customer
def reports(request, tracking_code):
    sent_messages = OutgoingDone.objects.filter(delivery_status__icontains='DeliveredTo', track_code=tracking_code).count()
    messages_not_sent = OutgoingDone.objects.filter(Q(delivery_status__icontains="InvalidMsisdn")|
                                                    Q(delivery_status__icontains="DeliveryImpossible")|
                                                    Q(delivery_status__icontains="AbsentSubscriber")|
                                                    Q(delivery_status__icontains="SenderName Blacklisted"),
                                                    track_code=tracking_code).count()
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        context = {
            'sent_messages': sent_messages,
            'messages_not_sent': messages_not_sent,
            'airtel_messages': OutgoingDone.objects.filter(customer_id=customer.id, track_code=tracking_code, delivery_status__contains='success').count()
        }
        return render(request, 'sms/reports.html', context)
    else:
        sub_acc = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first()
        customer = sub_acc.owner
        context = {
            'sent_messages': sent_messages,
            'messages_not_sent': messages_not_sent,
            'airtel_messages': OutgoingDone.objects.filter(customer_id=customer.id, track_code=tracking_code,
                                                           delivery_status__contains='success').count()
        }
        return render(request, 'sms/reports.html', context)


@login_required()
@is_user_customer
def customer_contacts(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer.id)

        context = {
            'customer': customer,
            'groups': groups
        }
        return render(request, 'sms/contacts.html', context)


def roberms_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            if Customer.objects.filter(user_ptr_id=user.id).count() > 0:
                return redirect('sms:apps')
            elif SalesPerson.objects.filter(user_ptr_id=user.id).count() > 0:
                return redirect('salesperson:dashboard')
            elif Manager.objects.filter(user_ptr_id=user.id).count() > 0:
                return redirect('roberms_admin:dashboard')
            elif CustomerSubAccounts.objects.filter(user_ptr_id=user.id).count() > 0:
                print(user.id)
                return redirect('sms:apps')
        else:
            messages.error(request, 'Invalid Email Or Password')
            return redirect('sms:login')
    return render(request, 'registration/login.html')


@login_required()
@is_user_customer
def create_group(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        if request.method == "POST":
            Group.objects.create(
                customer=customer,
                name=request.POST['name']
            )
            return redirect('sms:customer_contacts')
        return render(request, 'group/create_group.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        if request.method == "POST":
            Group.objects.create(
                customer=customer,
                name=request.POST['name']
            )
            return redirect('sms:customer_contacts')
        return render(request, 'group/create_group.html')


@login_required()
@is_user_customer
def group_contacts(request, group_id):
    contact_list = Contact.objects.filter(group_id=group_id)
    # paginator = Paginator(contact_list, 500)
    # page = request.GET.get('page', 1)

    # try:
    #     contacts = paginator.page(page)
    # except PageNotAnInteger:
    #     contacts = paginator.page(1)
    # except EmptyPage:
    #     contacts = paginator.page(paginator.num_pages)

    context = {
        'contacts': contact_list,
        'group': Group.objects.filter(id=group_id).first(),
        # 'record_count': paginator.count
    }
    return render(request, 'contacts/group_contacts.html', context)


@login_required()
@is_user_customer
def activate_deactivate_contact(request, contact_id):
    contact = Contact.objects.get(id=contact_id)
    if contact.is_active:
        contact.is_active = False
        contact.save()
        messages.success(request, 'Contact deactivated successfully')
        return redirect('sms:update_contact', contact.id)
    elif not contact.is_active:
        contact.is_active = True
        contact.save()
        messages.success(request, 'Contact activated successfully')
        return redirect('sms:update_contact', contact.id)


@login_required()
@is_user_customer
def create_contact(request, group_id):
    group = Group.objects.filter(id=group_id).first()
    if request.method == 'POST':
        phone_number = f"{254}{request.POST['phone_number'].replace(' ', '')[-9:]}"
        Contact.objects.update_or_create(
            group_id=group_id,
            name=request.POST['name'],
            email=request.POST['email'],
            phone_number=phone_number
        )
        return redirect('sms:sample_datatable', group_id)
    context = {
        'group': group
    }
    return render(request, 'contacts/create_contact.html', context)

@timed
@login_required()
@is_user_customer
def import_contacts(request, group_id):
    if request.method == 'POST':
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)


@task()
def store_contact_task(group_id, extension, uploaded_file_url, f_path):
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        with open(file_path, 'r') as f:
            firstline = True
            for row in csv.reader(f):
                if firstline:
                    firstline = False
                    continue
                else:
                    # print(row[2])
                    p = f"{254}{row[1].replace(' ', '')[-9:]}"
                    Contact.objects.update_or_create(
                        name=row[0],
                        group_id=group_id,
                        phone_number=int(p),
                        email=row[2]
                    )
            CustomerTask.objects.filter(task_id=store_contact_task.id).update(
                status_complete=True
            )
            return 'completed'
    else:
        # print('work')
        # print(file_path)
        workbook = load_workbook(filename=f_path, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        for i in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=i, column=2).value != '':
                group_id = group_id
                name = worksheet.cell(row=i, column=1).value
                phone_number = str(worksheet.cell(row=i, column=2).value)
                email = worksheet.cell(row=i, column=3).value
                p = f"{254}{phone_number.replace(' ', '')[-9:]}"
                try:
                    Contact.objects.update_or_create(
                        name=name,
                        group_id=group_id,
                        phone_number=int(p),
                        email=email
                    )
                except TypeError:
                    continue
                except ValueError:
                    continue
                print('saved')
        # CustomerTask.objects.filter(task_id=store_contact_task.id).update(
        #     status_complete=True
        # )
        return 'completed'


@timed
@login_required()
@is_user_customer
def express_import_contacts(request, group_id):
    if request.method == 'POST':
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = new_store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            file = request.FILES['my_file']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            uploaded_file_url = fs.url(filename)
            f_path = uploaded_file_url.split('/', 1)[1]
            extension = file.name.rsplit('.', 1)[1]
            s = new_store_contact_task.delay(group_id, extension, uploaded_file_url, f_path)
            CustomerTask.objects.create(
                customer=customer,
                task_id=s.id
            )
            return redirect('sms:contacts_upload_status', s.id)


@task()
def new_store_contact_task(group_id, extension, uploaded_file_url, f_path):
    if extension == 'csv':
        file_path = uploaded_file_url.split('/', 1)[1]
        with open(file_path, 'r') as f:
            firstline = True
            contacts = []
            for row in csv.reader(f):
                if firstline:
                    firstline = False
                    continue
                else:
                    # print(row[2])
                    p = f"{254}{row[1].replace(' ', '')[-9:]}"
                    contact = Contact(
                        name=str(row[0]),
                        group_id=group_id,
                        phone_number=int(p),
                        email=row[2]
                    )
                    contacts.append(contact)
                    if len(contacts) >= 20000:
                        Contact.objects.bulk_create(contacts)
                        contacts.clear()
            Contact.objects.bulk_create(contacts)
            return 'completed'
    else:
        # print('work')
        # print(file_path)
        workbook = load_workbook(filename=f_path, read_only=False)
        worksheet = workbook[workbook.sheetnames[0]]
        contacts = []
        for i in range(2, worksheet.max_row):
            # if worksheet.cell(row=i, column=2).value != '':
            group_id = group_id
            name = worksheet.cell(row=i, column=1).value
            phone_number = str(worksheet.cell(row=i, column=2).value)
            email = worksheet.cell(row=i, column=3).value
            p = f"{254}{phone_number.replace(' ', '')[-9:]}"

            # Contact.objects.update_or_create(
            #     name=name,
            #     group_id=group_id,
            #     phone_number=int(p),
            #     email=email
            # )
            contact = Contact(
                name=name,
                group_id=group_id,
                phone_number=int(p),
                email=email
            )
            contacts.append(contact)
            if len(contacts) >= 20000:
                Contact.objects.bulk_create(contacts)
                contacts.clear()
            else:
                continue
        Contact.objects.bulk_create(contacts)
        return 'completed'


@login_required()
@is_user_customer
def contacts_upload_status(request, task_id):
    context = {
        'task_id': task_id
    }
    return render(request, 'contacts/upload_status.html', context)


def poll_contact_upload_state(request, task_id):
    """ A view to report the progress to the user """

    job = AsyncResult(task_id)
    data = job.result or job.state
    response_data = {
        'state': job.state,
        'details': job.result,
    }
    return HttpResponse(json.dumps(response_data), content_type='application/json')


@login_required()
@is_user_customer
def comprehensive_reports(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        # messages_sent = OutgoingDone.objects.filter(customer_id=customer.id).values('track_code', 'sent_time').distinct()
        messages_sent = OutgoingDone.objects.raw(
            f'SELECT * FROM sms_outgoingdone WHERE customer_id={customer.id} GROUP BY track_code'
        )
        track_codes = []
        # for message in messages_sent.iterator():
        #     track_codes.append(message.track_code)

        print(messages_sent)
        context = {
            'track_codes': messages_sent
        }
        return render(request, 'sms/comprehensive_reports.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        # messages_sent = OutgoingDone.objects.filter(customer_id=customer.id).values('track_code', 'sent_time').distinct()
        messages_sent = OutgoingDone.objects.raw(
            f'SELECT * FROM sms_outgoingdone WHERE customer_id={customer.id} GROUP BY track_code'
        )
        # track_codes = []
        # for message in messages_sent.iterator():
        #     track_codes.append(message.track_code)
        context = {
            'track_codes': messages_sent
        }
        return render(request, 'sms/comprehensive_reports.html', context)


@login_required()
@is_user_customer
def report_details(request, track_code):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        o_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id)
        sent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                    delivery_status__contains='DeliveredTo')
        unsent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id).exclude(
            delivery_status__contains='DeliveredTo')
        pending_delivery_status = OutgoingNew.objects.filter(track_code=track_code, customer_id=customer.id,
                                                             delivery_status__isnull=True)
        context = {
            # 'o_messages': o_messages,
            'sent_messages': sent_messages.count(),
            'unsent_messages': unsent_messages.count(),
            'pending_delivery_status': pending_delivery_status.count(),
            'track_code': track_code,
            'airtel_messages': OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                    delivery_status__contains='success').count()
        }
        return render(request, 'sms/report_details.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        # o_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id)
        sent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                    delivery_status__contains='DeliveredTo')
        unsent_messages = OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id).exclude(
            delivery_status__contains='DeliveredTo')
        pending_delivery_status = OutgoingNew.objects.filter(track_code=track_code, customer_id=customer.id,
                                                             delivery_status__isnull=True)

        context = {
            # 'o_messages': o_messages,
            'sent_messages': sent_messages.count(),
            'unsent_messages': unsent_messages.count(),
            'pending_delivery_status': pending_delivery_status.count(),
            'track_code': track_code,
            'airtel_messages': OutgoingDone.objects.filter(track_code=track_code, customer_id=customer.id,
                                                           delivery_status__contains='success').count()
        }
        return render(request, 'sms/report_details.html', context)


# @login_required()
# @is_user_customer
# def generate_sms_report(request, track_code):
#     messages = OutgoingDone.objects.filter(track_code=track_code)
#     customer = Customer.objects.filter(user_ptr_id=messages.first().customer_id).first()
#     time = datetime.datetime.now()
#     file_path = 'media/reports/%s' % (customer.sender_name.replace(" ", "_"))
#     filename = "%s_%d_%d.pdf" % (messages.first().track_code, time.year, time.month)
#     full_path = f"{file_path}/{filename}"
#     data = {
#         'today': datetime.datetime.today(),
#         'messages': messages
#     }
#
#     html_string = render_to_string('sms/sms_report.html', data)
#
#     html = HTML(string=html_string)
#     html.write_pdf(target=full_path)
#
#     fs = FileSystemStorage(file_path)
#     with fs.open(filename) as pdf:
#         print(pdf)
#         response = HttpResponse(pdf, content_type='application/pdf')
#         # response['Content-Disposition'] = f'attachment; filename="{filename}"'
#         return response


@login_required()
@is_user_customer
def personalized_sms_menu(request):
    return render(request, 'sms/personalized_sms_menu.html')


@login_required()
@is_user_customer
def personalized_from_contact_list(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        groups = Group.objects.filter(customer=customer)
        if request.method == 'POST':
            message = request.POST.get('Message')
            group_ids = [g.id for g in groups]
            if 'all_groups' in request.POST:
                context = {
                    'message': message,
                    'groups': group_ids
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
            else:
                context = {
                    'message': message,
                    'group': request.POST['group']
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
        context = {
            'groups': groups
        }
        return render(request, 'sms/personalized_from_contact_list.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        groups = Group.objects.filter(customer=customer)
        if request.method == 'POST':
            message = request.POST.get('Message')
            group_ids = [g.id for g in groups]
            if 'all_groups' in request.POST:
                context = {
                    'message': message,
                    'groups': group_ids
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
            else:
                context = {
                    'message': message,
                    'group': request.POST['group']
                }
                request.session['c_data'] = context
                return redirect('sms:c_sample_merged')
        context = {
            'groups': groups
        }
        return render(request, 'sms/personalized_from_contact_list.html', context)


def c_sample_merged(request):
    data = request.session.get('c_data')
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        parameters = get_message_parameters(message=data['message'])
        if 'group' in data:
            c_group = Group.objects.filter(id=data['group']).first()
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for contact in Contact.objects.filter(group_id=request.POST['group'], is_active=True):
                    complete_message = message
                    if parameters:
                        for parameter in parameters:
                            if parameter == 'Name':
                                complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                    else:
                        to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'group':
                        contacts = Contact.objects.filter(group_id=b)
                        for contact in contacts:
                            complete_message = data['message']
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': c_group.id
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)
        elif 'groups' in data:
            parameters = get_message_parameters(message=data['message'])
            groups = Group.objects.filter(id__in=data['groups'])

            if request.method == 'POST':
                print('got here')
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']
                print(groups)
                for group in groups:
                    for contact in Contact.objects.filter(group_id=group.id, is_active=True):
                        complete_message = message
                        if parameters:
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                        else:
                            to_be_sent[contact.phone_number] = complete_message
                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'groups':
                        for g in groups:
                            contacts = Contact.objects.filter(group_id=g.id)
                            for contact in contacts:
                                complete_message = data['message']
                                for parameter in parameters:
                                    if parameter == 'Name':
                                        complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': groups[0]
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        parameters = get_message_parameters(message=data['message'])
        if 'group' in data:
            c_group = Group.objects.filter(id=data['group']).first()
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for contact in Contact.objects.filter(group_id=request.POST['group'], is_active=True):
                    complete_message = message
                    if parameters:
                        for parameter in parameters:
                            if parameter == 'Name':
                                complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                    else:
                        to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'group':
                        contacts = Contact.objects.filter(group_id=b)
                        for contact in contacts:
                            complete_message = data['message']
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                            to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': c_group.id
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)
        elif 'groups' in data:
            groups = Group.objects.filter(id__in=data['groups'])
            if request.method == 'POST':
                trackingcode = random.randint(1, 1000000)
                while OutgoingNew.objects.filter(track_code=trackingcode).count() > 0:
                    trackingcode = random.randint(1, 1000000)

                to_be_sent = {}
                total_message_cost = 0
                message = data['message']

                for group in groups:
                    for contact in Contact.objects.filter(group_id=group.id, is_active=True):
                        complete_message = message
                        if parameters:
                            for parameter in parameters:
                                if parameter == 'Name':
                                    complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                        else:
                            to_be_sent[contact.phone_number] = complete_message

                for a, b in to_be_sent.items():
                    total_message_cost += calculate_message_cost(message=b)
                if customer.credit < total_message_cost:
                    messages.error(request,
                                   'You do not have enough credit to make this request, kindly recharge to proceed')
                    return render(request, 'sms/personalized_from_contact_list.html')
                else:
                    from_group_send.delay(customer.id, total_message_cost, to_be_sent, trackingcode)
                    return redirect('sms:customer_reports', trackingcode)
            else:
                to_be_sent = {}
                for a, b in data.items():
                    if a == 'groups':
                        for g in groups:
                            contacts = Contact.objects.filter(group_id=g.id)
                            for contact in contacts:
                                complete_message = data['message']
                                for parameter in parameters:
                                    if parameter == 'Name':
                                        complete_message = complete_message.replace('[Name]', contact.name)
                                to_be_sent[contact.phone_number] = complete_message
                context = {
                    'merged_sample_data': to_be_sent,
                    'message': data['message'].replace("\n", "<br>").replace("\r", " "),
                    'group': groups[0]
                }
                # pprint(context)
                return render(request, 'sms/c_sample_merged.html', context)


@task()
def from_group_send(customer_id, total_message_cost, to_be_sent, trackingcode):
    customer = Customer.objects.filter(id=customer_id).first()
    new_credit = customer.credit - total_message_cost
    customer.credit = new_credit
    customer.save()

    # messages = []
    for a, b in to_be_sent.items():
        outgoing_new, created = OutgoingNew.objects.update_or_create(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=a,
            text_message=b,
            track_code=trackingcode,
            sent_time=timezone.now()
        )

    #     messages.append(
    #         OutgoingNew(
    #             customer=customer,
    #             service_id=customer.service_id,
    #             access_code=customer.access_code,
    #             phone_number=a,
    #             text_message=b,
    #             track_code=trackingcode
    #         )
    #     )
    #     if len(messages) > 1000:
    #         OutgoingNew.objects.bulk_create(messages)
    # if len(messages) > 0:
    #     OutgoingNew.objects.bulk_create(messages)
    send_usage(total_message_cost)
    return 'completed insertion'


@login_required
def delete_group(request, group_id):
    group = Group.objects.filter(id=group_id).first()
    group.delete()
    messages.success(request, 'Group Successfully Deleted')
    return redirect("sms:customer_contacts")


@login_required
def delete_contact(request, contact_id):
    contact = Contact.objects.filter(id=contact_id).first()
    group = contact.group
    contact.delete()
    messages.success(request, 'Contact Successfully Deleted')
    return redirect("sms:sample_datatable", group.id)


def customer_top_up(request):
    tracking_code = random.randint(1, 1000000)
    sender_phone = request.POST['sender_phone']
    transaction_reference = request.POST['trans_id']
    amount = request.POST['amount']
    till_number = request.POST['till_number']
    timestamp = request.POST['transaction_timestamp']
    first_name = request.POST['first_name']
    last_name = request.POST['last_name']
    # transaction_type = request.POST['transaction_type']

    UserTopUp.objects.create(
        phone_number=sender_phone,
        transaction_ref=transaction_reference,
        amount=amount,
        till_number=till_number,
        f_name=first_name,
        l_name=last_name,
        verify_code=tracking_code,
        timestamp=timestamp
    )

    message = f"Thank you {first_name}" \
        "for paying for Roberms sms service. To automatically load the credit into your account, " \
        f"use this code to verify your payment. {tracking_code}"
    customer = Customer.objects.filter().first()
    if customer is not None:
        # if str.isdigit(customer.access_code):
        #     outgoing = Outgoing.objects.create(
        #         customer=customer,
        #         service_id=customer.service_id,
        #         access_code=customer.access_code,
        #         phone_number=sender_phone,
        #         text_message=message,
        #         track_code=tracking_code
        #     )
        # else:
        outgoing_new = OutgoingNew.objects.create(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=sender_phone,
            text_message=message,
            track_code=tracking_code,
            sent_time=timezone.now()
        )
    else:
        customer = CustomerSubAccounts.objects.filter().first().owner
        # if str.isdigit(customer.access_code):
        #     outgoing = Outgoing.objects.create(
        #         customer=customer,
        #         service_id=customer.service_id,
        #         access_code=customer.access_code,
        #         phone_number=sender_phone,
        #         text_message=message,
        #         track_code=tracking_code
        #     )
        # else:
        outgoing_new = OutgoingNew.objects.create(
            customer=customer,
            service_id=customer.service_id,
            access_code=customer.access_code,
            phone_number=sender_phone,
            text_message=message,
            track_code=tracking_code,
            sent_time=timezone.now()
        )
        # sdp = SDP()
        # response = sdp.send_sms_customized(service_id=outgoing.service_id, recipients=[outgoing.phone_number],
        #                                    message=outgoing.text_message, sender_code='')


@login_required()
@is_user_customer
def verify_payment(request):
    if request.method == 'POST':
        user_top_up = Sms_TopUp.objects.filter(verifycode=request.POST['verification_code'], verified=0).first()
        if user_top_up:
            credit = 0
            if float(user_top_up.amount) < 500:
                credit = float(user_top_up.amount) * float(1)
            elif 500 <= float(user_top_up.amount) <= float(9999):
                credit = float(user_top_up.amount) * float(1.67)
            elif 10000 <= float(user_top_up.amount) <= float(49999):
                credit = float(user_top_up.amount) * float(2)
            elif float(user_top_up.amount) >= float(50000):
                credit = float(user_top_up.amount) * float(2.5)

            customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
            if customer is not None:
                existing_credit = customer.credit
                customer.credit = existing_credit + credit
                user_top_up.verified = 1
                user_top_up.user_id = customer.id
                user_top_up.created_at = datetime.datetime.now()
                user_top_up.save()
                customer.save()
                ManagerTopUp.objects.create(
                    sms_count=credit,
                    amount=user_top_up.amount,
                    user_id=customer.id,
                    timestamp=datetime.datetime.now()
                )
                messages.success(request, 'Top Up Successful')
                return redirect('sms:apps')
            else:
                customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
                existing_credit = customer.credit
                customer.credit = existing_credit + credit
                user_top_up.verified = 1
                user_top_up.user_id = customer.id
                user_top_up.created_at = datetime.datetime.now()
                user_top_up.save()
                customer.save()
                ManagerTopUp.objects.create(
                    sms_count=credit,
                    amount=user_top_up.amount,
                    user_id=customer.id,
                    timestamp=datetime.datetime.now()
                )
                messages.success(request, 'Top Up Successful')
                return redirect('sms:apps')
        else:
            messages.error(request, 'The Verification Code You Entered is Invalid')
            return redirect('sms:verify_payment')
    return render(request, 'top_up/payment_verification_code.html')


@login_required
def update_group(request, group_id):
    group = Group.objects.filter(id=group_id).first()
    if request.method == 'POST':
        Group.objects.filter(id=group_id).update(
            name=request.POST['name']
        )
        messages.success(request, 'Group Successfully Updated')
        return redirect('sms:customer_contacts')
    context = {
        'group':  group
    }
    return render(request, 'sms/update_group.html', context)


@login_required
def update_contact(request, contact_id):
    contact = Contact.objects.filter(id=contact_id).first()
    if request.method == 'POST':
        Contact.objects.filter(id=contact_id).update(
            name=request.POST['name'],
            email=request.POST['email'],
            phone_number=request.POST['phone_number']
        )
        messages.success(request, 'Contact Successfully Updated')
        return redirect('sms:group_contacts', contact.group.id)
    context = {
        'group': contact.group,
        'contact': contact
    }
    return render(request, 'sms/update_contact.html', context)


@login_required
def sms_reports(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        m = OutgoingDone.objects.filter(customer_id=customer.id)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/sms_reports.html', context)
        else:
            return render(request, 'sms/sms_reports.html')
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        m = OutgoingDone.objects.filter(customer_id=customer.id)
        if m.count() > 0:
            # pprint(m)
            context = {
                'outgoings': m
            }
            return render(request, 'sms/sms_reports.html', context)
        else:
            return render(request, 'sms/sms_reports.html')


@login_required()
def customer_till_numbers(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        till_numbers = Till_Numbers.objects.filter(customer=customer)
        context = {
            'till_numbers': till_numbers
        }
        return render(request, 'sms/customer_till_numbers.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        till_numbers = Till_Numbers.objects.filter(customer=customer)
        context = {
            'till_numbers': till_numbers
        }
        return render(request, 'sms/customer_till_numbers.html', context)


@login_required()
def delete_till_number(request, till_number_id):
    till_number = Till_Numbers.objects.filter(id=till_number_id).first()
    if till_number is not None:
        till_number.delete()
        messages.success(request, 'Till Number Deleted Successfully')
        return redirect('sms:customer_till_numbers')
    else:
        messages.success(request, 'Till Number Does Not Exist')
        return redirect('sms:customer_till_numbers')


@login_required()
def edit_till_number(request, till_number_id):
    till_number = Till_Numbers.objects.get(id=till_number_id)
    if request.method == 'POST':
        till_number = request.POST['till_number']
        message = request.POST['message']

        Till_Numbers.objects.filter(id=till_number_id).update(
            till=till_number,
            message=message
        )

        return redirect('sms:customer_till_numbers')
    else:
        context = {
            'till_number': till_number
        }
        return render(request, 'till_numbers/edit_till_number.html', context)

#
@login_required()
def add_till_number(request):
    if request.method == 'POST':
        till_number = request.POST['till_number']
        message = request.POST['message']
        customer = customer=Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            Till_Numbers.objects.create(
                customer=customer,
                till=till_number,
                message=message
            )
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            if customer is not None:
                Till_Numbers.objects.create(
                    customer=customer,
                    till=till_number,
                    message=message
                )
        return redirect('sms:customer_till_numbers')
    return render(request, 'till_numbers/add_till_number.html')


@login_required()
def delete_till_number(request, till_number_id):
    till = Till_Numbers.objects.filter(id=till_number_id).first()
    if till is not None:
        till.delete()
        messages.success(request, 'Till Number Deleted')
        return redirect('sms:customer_till_numbers')
    messages.error('Unable to delete till number')
    return redirect('sms:customer_till_numbers')


@login_required()
def credit_used(request):
    months = get_last_n_months(6)
    monthly_credit = []
    for month in list(set(months)):
        messages = OutgoingDone.objects.filter(sent_time__month=month)
        credit = 0
        for message in messages:
            credit += calculate_message_cost(message.text_message)
        monthly_credit.append(credit)

    context = {
        'months': list(set(months)),
        'monthly_credit': monthly_credit,
    }
    return render(request, 'sms/credit_usage.html', context)


@login_required()
def applications(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        applications = SenderNameApplication.objects.filter(customer_id=customer.id)
        context = {
            'applications': applications
        }
        return render(request, 'sender_name/sender_name_list.html', context)
    else:
        sub_account = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first()
        customer = sub_account.owner
        applications = SenderNameApplication.objects.filter(customer_id=customer.id)
        context = {
            'applications': applications
        }
        return render(request, 'sender_name/sender_name_list.html', context)


@login_required()
def new_application(request):
    if request.method == 'POST':
        date = datetime.datetime.today().date()
        sender_name = request.POST['sender_name']
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            n_application = SenderNameApplication.objects.create(
                customer=customer,
                sender_name=sender_name,
                application_date=date,
            )
            return redirect('sms:application_contacts', n_application.id)
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            n_application = SenderNameApplication.objects.create(
                customer=customer,
                sender_name=sender_name,
                application_date=date,
            )
            return redirect('sms:application_contacts', n_application.id)
    else:
        return render(request, 'sender_name/application_form.html')


@login_required()
def application_contacts(request, application_id):
    application_contacts = ApplicationContact.objects.filter(application_id=application_id)
    context = {
        'contacts': application_contacts,
        'application_id': application_id
    }
    return render(request, 'sender_name/contacts.html', context)


@login_required()
def add_application_contacts(request, application_id):
    application = SenderNameApplication.objects.filter(id=application_id).first()
    if request.method == 'POST':
        name = request.POST['name']
        phone_number = request.POST['phone_number']
        application_id = request.POST['application_id']
        ApplicationContact.objects.create(
            name=name,
            phone_number=phone_number,
            application_id=application_id
        )
        return redirect('sms:application_contacts', application_id)
    else:
        return render(request, 'sender_name/add_contacts.html', {'application':application})


@login_required()
def show_pdf(request, application_id):
    application = SenderNameApplication.objects.filter(id=application_id).first()
    application_contacts = ApplicationContact.objects.filter(application=application)
    data = {
        'application': application,
        'contacts': application_contacts
    }
    pdf = render_to_pdf('sms/sender_name_application.html', data)
    return HttpResponse(pdf, content_type='application/pdf')


# @login_required()
# def inbox(request):
#     customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
#     if customer is not None:
#         inbox_messages = Inbox.objects.filter(customer_id=customer.id)
#         context = {
#             'messages': inbox_messages
#         }
#         return render(request, 'inbox/general_inbox.html', context)
#     else:
#         customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
#         inbox_messages = Inbox.objects.filter(customer_id=customer.id)
#         context = {
#             'messages': inbox_messages
#         }
#         return render(request, 'inbox/general_inbox.html', context)


def airtel_callback(request):
    print(request)
    return


def trial(request):
    print(get_access_token())
    return HttpResponse(get_access_token())


def documentation(request):
    return render(request, 'api/documentation.html')


def offers(request):

    return render(request, 'sms/offers.html')


@login_required()
def new_tag(request):
    if request.method == "POST":
        tag = request.POST['tag']
        response = request.POST['default_response']
        customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
        if customer is not None:
            tag = f'#{tag}'
            tag = Tag.objects.update_or_create(
                hashtag=tag, customer=customer, defaults={'response': response}
            )
            messages.success(request, 'Tag created successfully')
            return redirect('sms:get_tags')
        else:
            tag = f'#{tag}'
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
            tag, created = Tag.objects.update_or_create(
                hashtag=tag, customer=customer, defaults={'response': response}
            )
            messages.success(request, 'Tag created successfully')
            return redirect('sms:get_tags')
    else:
        return render(request, 'inbox/new_tag.html')


def get_tags(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    if customer is not None:
        context = {
            'tags': Tag.objects.filter(customer_id=customer.id)
        }
        return render(request, 'inbox/tags.html', context)
    else:
        customer = CustomerSubAccounts.objects.filter(user_ptr_id=request.user.id).first().owner
        context = {
            'tags': Tag.objects.filter(customer_id=customer.id)
        }
        return render(request, 'inbox/tags.html', context)


def tag_messages(request, tag_id):
    context = {
        'tag_messages': Inbox.objects.filter(tag_id=tag_id)
    }
    return render(request, 'inbox/general_inbox.html', context)


def temp_clean_sys(request):
    outgoings = OutgoingDone.objects.filter(
        Q(delivery_status="SenderName Blacklisted")|
        Q(delivery_status="InvalidMsisdn")|
        Q(delivery_status="AbsentSubscriber")
        ,
        Q(access_code="TREASURE")|
        Q(access_code="Nyumbani"),
    )
    print(outgoings.query)

    groups = Group.objects.filter(
        Q(customer_id=148) |
        Q(customer_id=52)
    )

    for outgoing in outgoings:
        for group in groups:
            contact = Contact.objects.filter(phone_number=outgoing.phone_number, group_id=group.id).first()
            if contact is not None:
                contact.is_active = False
                contact.save()
    return redirect("sms:apps")


from django_datatables_view.base_datatable_view import BaseDatatableView
from django.utils.html import escape


class OrderListJson(BaseDatatableView):
    model = Contact
    columns = ['id', 'name', 'phone_number', 'email', 'creation_date', 'is_active', 'actions']
    order_columns = ['id', 'name', 'phone_number', 'email', 'creation_date', '']
    max_display_length = 200

    def get_initial_queryset(self):
        print(self.kwargs['group_id'])
        return self.model.objects.filter(group_id=self.kwargs['group_id']).\
            extra(select={'creation_date':'DATE_FORMAT(created_at, "%%Y-%%m-%%d")'})

    def render_column(self, row, column):
        if column == 'user':
            return escape('{0} {1}'.format(row.customer_firstname, row.customer_lastname))
        else:
            return super(OrderListJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        search = self.request.GET.get('search[value]', None)
        print(self.kwargs['group_id'])
        if search:
            qs = qs.filter(Q(name__istartswith=search)|Q(phone_number__istartswith=search), group_id=self.kwargs['group_id'])

        filter_customer = self.request.GET.get('customer', None)
        # print(filter_customer)
        if filter_customer:
            customer_parts = filter_customer.split(' ')
            qs_params = None
            for part in customer_parts:
                q = Q(name__istartswith=part)|Q(phone_number__istartswith=part)
                qs_params = qs_params | q if qs_params else q
            qs = qs.filter(qs_params)
        return qs


@login_required()
def sample_datatable(request, group_id):
    context = {
        'group': Group.objects.get(id=group_id)
    }
    return render(request, 'contacts/sample.html', context)


class OrderReportJson(BaseDatatableView):
    model = OutgoingDone
    columns = ['text_message', 'phone_number', 'delivery_status', 'sent_time']
    order_columns = ['text_message', 'phone_number', 'delivery_status', 'sent_time']
    max_display_length = 200

    def get_initial_queryset(self):
        all_done = self.model.objects.filter(track_code=self.kwargs['track_code'])
        return all_done

    def render_column(self, row, column):
        return super(OrderReportJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        customer = Customer.objects.filter(user_ptr_id=self.request.user.id).first()
        if customer is None:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=self.request.user.id).first().owner
        print(customer.id)
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(text_message__istartswith=search)|
                           Q(phone_number__istartswith=search)|
                           Q(delivery_status__icontains=search), customer_id=customer.id)
        return qs


class AllMessagesJson(BaseDatatableView):
    model = OutgoingDone
    columns = ['phone_number', 'text_message','sent_time', 'delivery_status']
    order_columns = ['text_message', 'phone_number', 'delivery_status', 'sent_time']
    max_display_length = 200

    def get_initial_queryset(self):
        customer = Customer.objects.filter(user_ptr_id=self.request.user.id).first()
        if customer is not None:
            all_done = self.model.objects.filter(customer_id=customer.id)
            return all_done
        else:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=self.request.user.id).first().owner
            all_done = self.model.objects.filter(customer_id=customer.id)
            return all_done

    def render_column(self, row, column):
        return super(AllMessagesJson, self).render_column(row, column)

    def filter_queryset(self, qs):
        customer = Customer.objects.filter(user_ptr_id=self.request.user.id).first()
        if customer is None:
            customer = CustomerSubAccounts.objects.filter(user_ptr_id=self.request.user.id).first().owner
        print(customer.id)
        search = self.request.GET.get('search[value]', None)
        if search:
            qs = qs.filter(Q(text_message__istartswith=search)|
                           Q(phone_number__istartswith=search)|
                           Q(delivery_status__icontains=search), customer_id=customer.id)
        return qs


@login_required()
@is_user_customer
def generate_sms_report(request, track_code):
    messages = OutgoingDone.objects.filter(track_code=track_code)
    customer = Customer.objects.filter(user_ptr_id=messages.first().customer_id).first()
    time = datetime.datetime.now()
    file_path = 'media/reports/%s' % (customer.sender_name.replace(" ", "_"))
    filename = "%s_%d_%d.xlsx" % (messages.first().track_code, time.year, time.month)
    full_path = f"{file_path}/{filename}"
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook = Workbook()
    summary_sheet = workbook.get_sheet_by_name('Sheet')
    summary_sheet.title = 'Report'
    cell = summary_sheet.cell(row=1, column=1)
    cell.value = 'Message Report'
    cell.alignment = Alignment(horizontal='center', vertical='center')
    summary_sheet.merge_cells('A1:E1')
    summary_sheet.append((' ', 'TEXT MESSAGE', 'MSISDN', 'DELIVERY STATUS', 'SENT DATE'))

    number = 1
    for message in messages:
        summary_sheet.append((number, message.text_message, message.phone_number, message.delivery_status, message.sent_time))
        number += 1

    workbook.save(full_path)
    context = {
        'file_path': full_path
    }
    return render(request, 'sms/download_pdf.html', context)


@login_required()
@is_user_customer
def my_payments(request):
    customer = Customer.objects.filter(user_ptr_id=request.user.id).first()
    payments = Mpesa.objects.filter(customer_id=customer.id)

    context = {
        "customer": customer,
        "payments": payments
    }
    return render(request, 'payments/my_payments.html', context)


@login_required()
@is_user_customer
def st_ann_add_patient(request):
    groups = Group.objects.filter(customer_id=342)
    customer = Customer.objects.get(id=342)
    track_code = random.randint(1, 1000000)
    if request.method == 'POST':
        phone_number = request.POST["phone_number"]
        phone_number = f"{254}{phone_number[-9:]}"
        contact = Contact.objects.filter(phone_number=phone_number, group_id=request.POST["group"])
        if contact.count() > 0:
            c = contact.first()
            StAnnPatients.objects.create(
                name=request.POST["patient_name"],
                phone_number=c.phone_number,
                group_id=request.POST["group"]
            )
            messages.success(request, "Patient Added Successfully")
            return redirect("sms:customer_contacts")
        else:
            Contact.objects.create(
                name=request.POST["patient_name"],
                phone_number=phone_number,
                group_id=request.POST["group"]
            )
            StAnnPatients.objects.create(
                name=request.POST["patient_name"],
                phone_number=phone_number,
                group_id=request.POST["group"]
            )
            messages.success(request, "Patient Added Successfully")
            return redirect("sms:customer_contacts")
    context = {
        "groups": groups
    }
    return render(request, "st_ann/add_patient.html", context)





